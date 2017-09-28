from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Border, Side, Alignment, PatternFill
from openpyxl.utils.cell import get_column_letter

import pdb
import re
from operator import itemgetter
mykeys = ('LEVEL 1', 'LEVEL 2', 'LEVEL 3', 'LEVEL 4', 'LEVEL 5', 'LEVEL 6', 'LEVEL 7')
tracMatrixfilename = 'matrixFile'
wb = load_workbook(filename = 'excelFile.xlsx')
wb.save('outputExcelFile.xlsx')
titleStyle = Font(size = 14, bold = True)
titleBorder = Border(bottom = Side(style = 'thick'))
legendBorder = Border(bottom = Side(style = 'thick'), top = Side(style = 'thick'), left = Side(style = 'thick'), right = Side(style = 'thick'))

#elements of L6 associated to L2
def SearchL2L6(L2_string):
    global wb    
    L2_2_L6_sh = wb['L2-L6']
    result = None
    for row in tuple(L2_2_L6_sh.rows)[1:]:
        L2_from_L6 = row[0].value
        if L2_string == L2_from_L6:
            print(row[2].value)
            result = row[2].value.split('-')
            result = result[0]
            break
        elif not (L2_from_L6):
            break
    return result

def IsDeleted(element):
    if (element.strip()).lower() == 'deleted':
        return True
    else:
        return False
		
def IsNotAvailable(element):
    if (element.strip()).lower() == 'na' or (element.strip()).lower() == 'n.a.':
        return True	
    else:
        return False

def IsOrphan(element):
    if element is None:
        return False
    if (element.strip()).lower() == 'orphan':
        return True	
    else:
        return False

def updateL1(l1):
    temp_str = l1.split('-')
    loc_str = None
    temp_num = re.findall('[0 1 2 3 4 5 6 7 8 9 0]*', temp_str[3])
    if int(temp_num[0])<10 and len(temp_str) > 2:    
        loc_str = temp_str[0] + '-'  + temp_str[1] + '-'+ temp_str[2] + '-' + '0' + temp_num[0] + temp_str[3][len(temp_num[0]):]
    if loc_str == None:
        return l1
    else:
        return loc_str

def updateL2AndL3(el):
    temp_str = el.split('.') #ex. B23.020 B7.010
    loc_str = None
    temp_num = re.findall('[0 1 2 3 4 5 6 7 8 9 0]*', temp_str[0])
    if int(temp_num[1])<10 and len(temp_str) > 0:    
        loc_str = el[0] + '0' + temp_num[1] +'.' + temp_str[1]
    if loc_str == None:
        return el
    else:
        return loc_str

def invertL2AndL3(element):
    loc_str = None
    if element is None:
        return loc_str
    temp_str = element.split('.') #ex. B23.020 B7.010
    temp_num = re.findall('[0 1 2 3 4 5 6 7 8 9 0]*', temp_str[0])
    if int(temp_num[1])<10 and len(temp_str) > 0:    
        loc_str = element[0] + str(int(temp_num[1])) +'.' + temp_str[1]
    if loc_str == None:
        return element
    else:
        return loc_str

def WriteL7():
    #levelsDownward
    levelsDownward = LoadFromFile(tracMatrixfilename)
    wb2 = load_workbook(filename = 'outputExcelFile.xlsx')
    result = None
    try:
        levelsDownward_sh = wb2["LEVELS Downward"]
    except:
        levelsDownward_sh = wb2.create_sheet(title="LEVELS Downward")
    try:
        L6_7_trac_sh = wb2["L3-L7"]
    except:
        L6_7_trac_sh = wb2.create_sheet(title="L3-L7")
    count = 0
    for ii, row in enumerate(tuple(levelsDownward_sh.rows)[5:]):
        if not(not((row[6].value))):
            continue
        #To search L2/L3 the string is converted back to the original (ex. from B07.010 to B7.010)
        L3_from_levelsDown = invertL2AndL3(row[2].value)
        
        for row_L6_7 in tuple(L6_7_trac_sh.rows)[2:]:
            if L3_from_levelsDown == row_L6_7[0].value:
                count += 1
                
                if row_L6_7[1].value is None:
                    continue
                elif not(row_L6_7[1].value.split('-')[0] is None):
                    row[6].value = row_L6_7[1].value.split('-')[0]
                    temp_dict = levelsDownward[ii]
                    temp_dict[mykeys[6]] = row_L6_7[1].value.split('-')[0]
                    levelsDownward[ii] = temp_dict
                else:
                    row[6].value = row_L6_7[1].value
                    temp_dict = levelsDownward[ii]
                    temp_dict[mykeys[6]] = row_L6_7[1].value
                    levelsDownward[ii] = temp_dict

                if count%100 == 0:
                    print(count)
    wb2.save('outputExcelFile.xlsx')
    Save2File(tracMatrixfilename, levelsDownward)


def CreateLevelsMatrix():
    global mykeys, wb
    L1_2_L2_sh = wb[wb.get_sheet_names()[0]]
    L2_2_L3_sh = wb[wb.get_sheet_names()[1]]
    L3_2_L4_sh = wb[wb.get_sheet_names()[2]]
    
    final_dict = dict()
    final_list = list()
    
    print(wb.get_sheet_names())
    l2_item_old ="This is not a string"
    # OBTAIN A DICTIONARY WITH THE NAME OF FILE AS KEYWORD
    
    for row in tuple(L1_2_L2_sh.rows)[1:29]:
        L2r = (row[2].value).split('\n')
        L1r = (row[0].value.split('\n'))
        
    	#for ordering        
        for l1_item in L1r: #should be 1
            if IsDeleted(row[2].value) or IsNotAvailable(row[2].value):
                final_dict = dict()
                final_dict = {mykeys[0] : updateL1(l1_item), mykeys[1] : row[2].value, mykeys[2] : row[2].value, mykeys[3] : row[2].value, mykeys[4] : row[2].value, mykeys[5] : row[2].value, mykeys[6] : row[2].value}
                final_list.append(final_dict)
            else:
                counter = 0
                for l2_item in L2r: #can be more than 1
                    for l3_item in tuple(L2_2_L3_sh.rows)[1:32]:
                        counter += 1
                        #pdb.set_trace()
                        if  l2_item.strip() == l3_item[0].value.strip(): #if the LEVEL2 name of the L1->L2 table corresponds to the one of L2->L3 table   							
                            l6 = ""
                            if IsDeleted(l3_item[1].value.strip()) or IsNotAvailable(l3_item[1].value.strip()):
                                final_dict = dict()
                                final_dict = {mykeys[0] : updateL1(l1_item), mykeys[1] : updateL2AndL3(l2_item), mykeys[2] : l3_item[1].value, mykeys[3] : l3_item[1].value, mykeys[4] : l3_item[1].value, mykeys[5] : l3_item[1].value, mykeys[6] : l3_item[1].value}
                                final_list.append(final_dict)
                            else:
                                for l4 in tuple(L3_2_L4_sh.rows)[1:]:
                                    l7 = ""
                                    if l2_item == l2_item_old:
                                        l6_item = l6_item_old
                                    else:
                                        l6_item = SearchL2L6(l2_item)
                                        l2_item_old = l2_item
                                        l6_item_old = l6_item
                                    if l4[0].value.strip() == l3_item[1].value.strip():
                                        final_dict = dict()
                                        final_dict = {mykeys[0] : updateL1(l1_item), mykeys[1] : updateL2AndL3(l2_item), mykeys[2] : updateL2AndL3(l4[0].value.strip()), mykeys[3] : l4[1].value.strip(), mykeys[4] : l4[2].value.strip(), mykeys[5] : l6_item, mykeys[6] : l7}
                                        final_list.append(final_dict)
                                        break
    
    levelsDownward = sorted(final_list, key=itemgetter(mykeys[0], mykeys[1], mykeys[2], mykeys[3], mykeys[4]))
    return levelsDownward

def Save2File(the_filename, my_list):
    import pickle
    with open(the_filename, 'wb') as f:
        pickle.dump(my_list, f)

def LoadFromFile(the_filename):
    import pickle
    with open(the_filename, 'rb') as f:
        my_list = pickle.load(f)
    return my_list

def UpdateFile():
    global mykeys, wb
    levelsDownward = CreateLevelsMatrix()
    Save2File(tracMatrixfilename, levelsDownward)	

def CreateLegend(ws):
    ws['A1'] = 'NOTES'
    ws['A1'].font = titleStyle #ws['A1'].style = titleStyle #old
    ws['A1'].border = titleBorder
    ws['B1'].border = titleBorder
    ws['C1'].border = titleBorder
    ws['D1'].border = titleBorder
    ws['E1'].border = titleBorder
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:L1')
    ws['A2'] = """Some text 1."""
    ws['A3'] = """Some text 2."""
    ws['A4'] = """Some text 3."""
    ws['A4'].border = titleBorder
    ws['B4'].border = titleBorder
    ws['C4'].border = titleBorder
    ws['D4'].border = titleBorder
    ws['E4'].border = titleBorder
    ws.merge_cells('A2:L2')
    ws.merge_cells('A3:L3')
    ws.merge_cells('A4:L4')
    for jj in ['1', '4', '5']:
        for ii in range(0,12):
            mystr =  get_column_letter(ii+1) + jj
            ws[mystr].border = titleBorder
	
def WriteLevelsDownward2XL():
    global mykeys, wb, titleStyle
    try:
        ws1 = wb["LEVELS Downward"]
    except:
        ws1 = wb.create_sheet("LEVELS Downward")
    
    ws1['A5'] = 'LEVEL 1'
    ws1['B5'] = 'LEVEL 2'
    ws1['C5'] = 'LEVEL 3'
    ws1['D5'] = 'LEVEL 4'
    ws1['E5'] = 'LEVEL 5'
    ws1['F5'] = 'LEVEL 6'
    ws1['G5'] = 'LEVEL 7'
    col_widths = [len(ws1['A5'].value)*2, len(ws1['B5'].value)*2, len(ws1['C5'].value)*2, len(ws1['D5'].value)*2, len(ws1['E5'].value)*2, len(ws1['F5'].value)*3, len(ws1['G5'].value)*3]
    
    ws1['A5'].font = titleStyle #ws1['A5'].style = titleStyle
    ws1['B5'].font = titleStyle#ws1['B5'].style = titleStyle
    ws1['C5'].font = titleStyle#ws1['C5'].style = titleStyle
    ws1['D5'].font = titleStyle#ws1['D5'].style = titleStyle
    ws1['E5'].font = titleStyle#ws1['E5'].style = titleStyle
    ws1['F5'].font = titleStyle#ws1['F5'].style = titleStyle
    ws1['G5'].font = titleStyle#ws1['G5'].style = titleStyle
    ws1['A5'].alignment = Alignment(horizontal='center')
    ws1['B5'].alignment = Alignment(horizontal='center')
    ws1['C5'].alignment = Alignment(horizontal='center')
    ws1['D5'].alignment = Alignment(horizontal='center')
    ws1['E5'].alignment = Alignment(horizontal='center')
    ws1['F5'].alignment = Alignment(horizontal='center')
    ws1['G5'].alignment = Alignment(horizontal='center')
    CreateLegend(ws1)
    CreateLegend(ws1)
    ws1.auto_filter.ref = "A5:G5"
    levelsDownward = LoadFromFile(tracMatrixfilename)
    count = 6
    col_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for item in levelsDownward:
        for ii in range(0, len(mykeys)):
            tempstr = col_list[ii] + str(count)
            ws1[tempstr] = item[mykeys[ii]]
            
            if len(item[mykeys[ii]]) > col_widths[ii]:
                col_widths[ii] = len(item[mykeys[ii]])
        count+= 1
    for i, column_width in enumerate(col_widths):
        ws1.column_dimensions[get_column_letter(i+1)].width = column_width    
    wb.save('outputExcelFile.xlsx')
    

def WriteLevelsUpward2XL():
    #Levels upward
    global mykeys, wb, titleStyle
    levelsDownward = LoadFromFile(tracMatrixfilename)
    levelsUpward = sorted(levelsDownward, key=itemgetter(mykeys[4], mykeys[3], mykeys[2], mykeys[1], mykeys[0]))
    try:
        ws2 = wb["LEVELS Upward Complete"]
    except:
        ws2 = wb.create_sheet("LEVELS Upward Complete")
    ws2['E5'] = 'L1'
    ws2['D5'] = 'L2'
    ws2['C5'] = 'L3'
    ws2['B5'] = 'L4'
    ws2['A5'] = 'L5'
    col_widths = [len(ws2['A5'].value)*2, len(ws2['B5'].value)*2, len(ws2['C5'].value)*2, len(ws2['D5'].value)*2, len(ws2['E5'].value)*2]
    ws2['A5'].font = titleStyle
    ws2['B5'].font = titleStyle
    ws2['C5'].font = titleStyle
    ws2['D5'].font = titleStyle
    ws2['E5'].font = titleStyle
    ws2['A5'].alignment = Alignment(horizontal='center')
    ws2['B5'].alignment = Alignment(horizontal='center')
    ws2['C5'].alignment = Alignment(horizontal='center')
    ws2['D5'].alignment = Alignment(horizontal='center')
    ws2['E5'].alignment = Alignment(horizontal='center')
    CreateLegend(ws2)
    ws2.auto_filter.ref = "A5:E5"
    count = 6
    col_list = ['A', 'B', 'C', 'D', 'E']
    col_list = list(reversed(col_list))
    col_widths = list(reversed(col_widths))
    for item in levelsUpward:
        for ii in reversed(range(0, len(mykeys)-2)):
            tempstr = col_list[ii] + str(count)
            ws2[tempstr] = item[mykeys[ii]]
            if len(item[mykeys[ii]]) > col_widths[ii]:
                col_widths[ii] = len(item[mykeys[ii]])
        count+= 1
    for i, column_width in enumerate(list(reversed(col_widths))):
        ws2.column_dimensions[get_column_letter(i+1)].width = column_width
    wb.save('outputExcelFile.xlsx')
	
def WriteL3Upward2XL():
    global mykeys, wb, titleStyle
    levelsDownward = LoadFromFile(tracMatrixfilename)
    l3upward = sorted(levelsDownward, key=itemgetter(mykeys[2], mykeys[1], mykeys[0], mykeys[1], mykeys[0]))
    try:
        ws = wb["LEVELS Upward"]
    except:
        ws = wb.create_sheet(title="LEVELS Upward")
    ws['A5'] = 'L3'
    ws['B5'] = 'L2'
    ws['C5'] = 'L1'
    col_widths = [len(ws['A5'].value)*2, len(ws['B5'].value)*2, len(ws['C5'].value)*2]

    ws['A5'].font = titleStyle
    ws['B5'].font = titleStyle
    ws['C5'].font = titleStyle
    ws['A5'].alignment = Alignment(horizontal='center')
    ws['B5'].alignment = Alignment(horizontal='center')
    ws['C5'].alignment = Alignment(horizontal='center')
    CreateLegend(ws)
    ws.auto_filter.ref = "A5:C5"
    count = 6
    col_list = ['C', 'B', 'A']
    for item in l3upward:
        for ii in range(0, len(col_list)):
            tempstr = col_list[ii] + str(count)
            ws[tempstr] = item[mykeys[ii]]
            if len(item[mykeys[ii]]) > col_widths[ii]:
                col_widths[ii] = len(item[mykeys[ii]])
        count +=1
    for i, column_width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width
    wb.save('outputExcelFile.xlsx')
       
def analyzeOrphanL2():
    global mykeys
    levelsDownward = LoadFromFile(tracMatrixfilename)
    wb = load_workbook(filename = 'outputExcelFile.xlsx')
    L2_2_L1_sh = wb['L2-L1']
    L2_2_L3_sh = wb['L2-L3']
    L3_2_L4_sh = wb['L3-L4-L5']
    l2_item_old ="This is not a string"
    l7_item = ""
    for l2_2_l1_row in tuple(L2_2_L1_sh.rows)[1:]:
        if IsOrphan(l2_2_l1_row[1].value):
            l6_item = ""
            for l2_2_l3_row in tuple(L2_2_L3_sh.rows)[1:]:
                if l2_2_l1_row[0].value == l2_2_l3_row[0].value:
                    if IsDeleted(l2_2_l3_row[1].value) or IsNotAvailable(l2_2_l3_row[1].value):
                        final_dict = dict()
                        final_dict = {mykeys[0] : 'ORPHAN', mykeys[1] : updateL2AndL3(l2_2_l3_row[0].value), mykeys[2] : l2_2_l3_row[1].value, mykeys[3] : l2_2_l3_row[1].value, mykeys[4] : l2_2_l3_row[1].value, mykeys[5] : l2_2_l3_row[1].value, mykeys[6] : l2_2_l3_row[1].value}
                        levelsDownward.append(final_dict)
                    else:                    
                        l2_item = l2_2_l3_row[0].value
                        if l2_item == l2_item_old:
                            l6_item = l6_item_old
                        else:
                            l6_item = SearchL2L6(l2_item)
                            l2_item_old = l2_item
                            l6_item_old = l6_item
                        for l4 in tuple(L3_2_L4_sh.rows)[1:]:
                            if l4[0].value.strip == l2_2_l3_row[1].value.strip:
                                final_dict = dict()
                                final_dict = {mykeys[0] : 'ORPHAN', mykeys[1] : updateL2AndL3(l2_2_l3_row[0].value), mykeys[2] : updateL2AndL3(l4[0].value.strip()), mykeys[3] : l4[1].value.strip(), mykeys[4] : l4[2].value.strip(), mykeys[5] : l6_item, mykeys[6] : l7_item}
                                levelsDownward.append(final_dict)
                                break
    Save2File(tracMatrixfilename, levelsDownward)
    
def analyzeOrphanL3():
    global mykeys
    levelsMatrix = LoadFromFile(tracMatrixfilename)
    wb = load_workbook(filename = 'outputExcelFile.xlsx')
    L3_2_L2_sh = wb['L3-L2']
    L3_2_L4_sh = wb['L3-L4-L5']
    l7_item = ""
    for l3_2_l2_row in tuple(L3_2_L2_sh.rows)[1:]:
        if IsOrphan(l3_2_l2_row[1].value):
            for l4 in tuple(L3_2_L4_sh.rows)[1:32]:
                if l4[0].value.strip == l3_2_l2_row[0].value.strip:
                    final_dict = dict()
                    final_dict = {mykeys[0] : 'ORPHAN', mykeys[1] : 'ORPHAN', mykeys[2] : updateL2AndL3(l4[0].value.strip()), mykeys[3] : l4[1].value.strip(), mykeys[4] : l4[2].value.strip(), mykeys[5] : 'ORPHAN', mykeys[6] : l7_item}
                    levelsMatrix.append(final_dict)
                    break
    Save2File(tracMatrixfilename, levelsMatrix)	


def main():
    UpdateFile()
    print("Matrix created")
    print("Adding orphans L2 elements")
    analyzeOrphanL2();
    print("L2 orphans elements")
    print("Adding orphans L3 elements")
    analyzeOrphanL3()
    analyzeOrphanL3()
    WriteLevelsDownward2XL()
    print("Downward-levels written to file")
    #UPWARD TRACEABILITY
    WriteLevelsUpward2XL()
    print("Upward-levels written to file")
    WriteL3Upward2XL()
    print("Matrix L2-upward written to file")
    WriteL7()
    print("L3 written to file")

if __name__ == '__main__':
    main()
